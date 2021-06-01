from tkinter import *
from tkinter import ttk, filedialog, messagebox
from main import *
import openpyxl
import uuid
import random


def clean_res():  # очистка полей значений
    XList.delete(0, END)
    YList.delete(0, END)


def save_file():  # сохранение файла
    if '.xlsx' not in OutStr.get():
        lace = filedialog.asksaveasfilename(defaultextension='.xlsx')
        OutStr.set(lace)
    else:
        lace = OutStr.get()

    l = [x.encode('utf8') for x in XList.get(0, END)]
    md4_list = [str(MD4(message).hexdigest()) for message in l]
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'Лист1'
    sheet.cell(row=1, column=1).value = 'Фраза'
    sheet.cell(row=1, column=2).value = 'Хэш'
    for i in range(len(md4_list)):
        sheet.cell(row=i + 2, column=1).value = l[i]
        sheet.cell(row=i + 2, column=2).value = md4_list[i]
    print('Документ успешно сформирован')
    wb.save(lace)
    print(l)


def load_file():  # загрузка из файла
    if '.xlsx' not in InStr.get():
        lace = filedialog.askopenfilename(defaultextension='.xlsx')
        InStr.set(lace)
    else:
        lace = InStr.get()

    for i in [(word, MD4(word.encode()).hexdigest()) for word in list_from_exel(lace[:-5], None)]:
        XList.insert(END, i[0])
        YList.insert(END, i[1])


def list_from_exel(file: str, paper: [str, None]) -> list:
    """
    From the file, where first column is Word make list of Words
    """
    if paper is None:
        paper = 'Лист1'
    l = []
    k = 2  # начальный строка

    # открытие для чтения
    wb = openpyxl.load_workbook(filename=file + '.xlsx', read_only=True)
    sheet = wb[paper]
    a = sheet.cell(row=k, column=1).value
    # пока не дойдёт до конца спика в файле добавляет значения
    while a is not None:
        l.append(a)
        k += 1
        a = sheet.cell(row=k, column=1).value
    return l


def md4():
    XList.insert(END, Word.get())
    YList.insert(END, str(MD4(Word.get().encode()).hexdigest()))


def rand_word():
    q = uuid.uuid4().__str__()
    q = q.__str__().encode()[:random.randint(0, len(q) - 1)]
    XList.insert(END, q)
    YList.insert(END, str(MD4(q).hexdigest()))


def lava():
    def text_to_bits(text):
        bits = bin(int.from_bytes(text, 'big'))[2:]
        return bits.zfill(8 * ((len(bits) + 7) // 8))

    def hex_to_bin(text):
        r = ''
        for i in text:
            q = bin(int(i, 16))[2:]
            while len(q) != 4:
                q = '0' + q
            r += q
        return r

    def text_from_bits(bits):
        n = int(bits, 2)
        return n.to_bytes((n.bit_length() + 7) // 8, 'big').decode()

    def change_bit(bits):
        r = bits[:5]
        r += '0' if bits[5] == '1' else '1'
        r += bits[6:]
        return r

    def diference(a, b):
        d = 0
        print(a.encode())
        print(b.encode())
        print(len(a.encode()))
        print(len(b.encode()))

        a = hex_to_bin(a)
        b = hex_to_bin(b)
        print(a)
        print(b)
        for i in range(len(a) - 1):
            if a[i] != b[i]:
                d += 1
        return d

    lace = filedialog.asksaveasfilename(defaultextension='.xlsx')
    if Count.get():
        count = int(Count.get())
    else:
        count = 15

    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'Лист1'
    sheet.cell(row=1, column=1).value = 'Фраза'
    sheet.cell(row=1, column=2).value = 'Хэш'
    sheet.cell(row=1, column=4).value = 'Фраза в 2 представлении'
    sheet.cell(row=1, column=5).value = 'Изменение случайного бита'
    sheet.cell(row=1, column=7).value = 'Хэш в 2'
    sheet.cell(row=1, column=8).value = 'Хэш в 2 измененного'
    sheet.cell(row=1, column=10).value = 'Битовое расстояние'

    for i in range(count):
        q = uuid.uuid4().__str__()
        text_in_bit = text_to_bits(q.encode())
        sheet.cell(row=i + 2, column=1).value = q
        sheet.cell(row=i + 2, column=2).value = MD4(q.__str__().encode()).hexdigest()
        sheet.cell(row=i + 2, column=4).value = text_in_bit
        sheet.cell(row=i + 2, column=5).value = change_bit(text_in_bit)
        sheet.cell(row=i + 2, column=7).value = hex_to_bin(MD4(q.__str__().encode()).hexdigest())
        sheet.cell(row=i + 2, column=8).value = hex_to_bin(
            MD4(text_from_bits(change_bit(text_in_bit)).encode()).hexdigest())
        sheet.cell(row=i + 2, column=10).value = diference(MD4(q.__str__().encode()).hexdigest(), MD4(
            text_from_bits(change_bit(text_in_bit)).encode()).hexdigest())
    wb.save(lace)
    print('готово')


def build(window):  # постройка окна
    global XList, YList

    style = ttk.Style()
    # available_themes = style.theme_names()
    style.theme_use('vista')  # задаём стиль по фану

    # создание элементов
    lbl = Label(text="Word = ").place(x=30, y=20)
    txt_A = Entry(width=20, textvariable=Word).place(x=100, y=20)
    lblB = Label(text="Кол-во фраз для лавинного эффекта = ").place(x=230, y=20)

    # ограничение на ввод
    def validate(new_value):  # +++
        return new_value == "" or new_value.isnumeric()

    vcmd = (window.register(validate), '%P')
    txt_B = Entry(window, width=20, textvariable=Count, validate='key', validatecommand=vcmd).place(x=500,
                                                                                                    y=20)

    Hex = Button(text="Хешировать", command=md4).place(x=30, y=50)

    lbl1 = Label(text="File IN = ").place(x=30, y=80)
    txt_I = Entry(width=20, textvariable=InStr).place(x=100, y=80)
    ButLoad1 = Button(text="Загрузить из файла", command=load_file).place(x=250, y=80)

    lbl2 = Label(text="File out = ").place(x=30, y=110)
    txt_O = Entry(width=20, textvariable=OutStr).place(x=100, y=110)
    ButSave = Button(text="Сохранить в файл", command=save_file).place(x=250, y=110)

    ButLoad = Button(text="Случайное значение", command=rand_word).place(x=150, y=50)
    ButLoad = Button(text="Лавинный эффект", command=lava).place(x=400, y=50)

    # ButQ = Button(ext='Выход', command=quit())

    lb31 = Label(text='Word ')
    lb32 = Label(text='Хэш')

    scrollbar1 = Scrollbar()
    scrollbar2 = Scrollbar()

    XList = Listbox(yscrollcommand=scrollbar1.set, bd=1)
    YList = Listbox(yscrollcommand=scrollbar2.set, bd=1)
    scrollbar1.config(command=XList.yview)  # привязка к листу
    scrollbar2.config(command=YList.yview)

    ButClean = Button(text='ОЧИСТИТЬ', command=clean_res)

    # разместить элементы
    lb31.place(x=40, y=130)
    lb32.place(x=180, y=130)

    XList.place(x=40, y=150, width=180, height=200)
    scrollbar1.place(x=220, y=150, height=200)
    YList.place(x=260, y=150, width=240, height=200)
    scrollbar2.place(x=500, y=150, height=200)

    ButClean.place(x=40, y=400, width=200, height=60)


if __name__ == '__main__':
    window = Tk()
    window.title("Добро пожаловать в Курсач")
    window.geometry('730x500')
    XList, YList = '', ''  # костыль, чтобы не делать через класс
    Word, Count, OutStr, InStr = (StringVar() for x in range(4))  # анологично
    build(window)
    window.mainloop()
